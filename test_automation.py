import openpyxl
from playwright.sync_api import sync_playwright
import os
import time

# === Configuration ===
EXCEL_FILE = "Assignment 1 - Test cases.xlsx"
SHEET_NAME = " Test cases"  # Note the leading space as found in the file
WEBSITE_URL = "https://www.pixelssuite.com/chat-translator"

def run_automation():
    """
    Enhanced automation script for Sinhala transliteration.
    Features dynamic waiting, reliable output capture, and Chrome integration.
    """
    # 1. Load Excel Workbook
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file '{EXCEL_FILE}' not found.")
        return

    print(f"Loading Excel file: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # Identify column indices
    headers = [cell.value for cell in ws[1]]
    try:
        input_col = headers.index("Input") + 1
        expected_col = headers.index("Expected output") + 1
        actual_col = headers.index("Actual output") + 1
        status_col = headers.index("Status") + 1
    except ValueError as e:
        print(f"Error: Missing column {e} in headers.")
        return

    # 2. Initialize Playwright
    with sync_playwright() as p:
        print("Launching browser...")
        # Use Chrome channel and non-headless mode for stability and visibility
        browser = p.chromium.launch(headless=False, channel="chrome")
        context = browser.new_context()
        page = context.new_page()

        print(f"Navigating to {WEBSITE_URL}...")
        try:
            page.goto(WEBSITE_URL, wait_until="networkidle", timeout=60000)
            page.wait_for_selector("textarea", timeout=20000)
        except Exception as e:
            print(f"Initial load failed: {e}")
            browser.close()
            return

        prev_output = "" # To ensure the output updates between rows

        # 3. Process Test Cases
        for row in range(2, ws.max_row + 1):
            input_text = ws.cell(row=row, column=input_col).value
            expected_output = ws.cell(row=row, column=expected_col).value

            if not input_text:
                continue

            print(f"\nProcessing row {row}: '{input_text}'")

            try:
                # 1. Input handling
                input_box = page.locator("textarea").nth(0)
                input_box.fill("") # Clear first
                input_box.fill(str(input_text)) # Use fill (not type)
                
                # 2. Button click
                trans_btn = page.locator("button:has-text('Transliterate')")
                if trans_btn.is_visible():
                    trans_btn.click()
                else:
                    # Fallback if button text is slightly different
                    page.get_by_role("button", name="Transliterate").click()

                # 3. Dynamic Waiting (IMPORTANT)
                # Wait until the output is not empty AND has changed from the previous value
                print("  Waiting for output to update...")
                try:
                    page.wait_for_function(
                        """(args) => {
                            const areas = document.querySelectorAll('textarea');
                            const target = areas.length > 1 ? areas[1] : document.querySelector('div[contenteditable="true"]');
                            if (!target) return false;
                            const current = (target.value || target.innerText || "").trim();
                            // Return true only if output is non-empty AND different from previous row's output
                            return current !== "" && current !== args.prev;
                        }""",
                        arg={"prev": prev_output},
                        timeout=10000 # Max wait 10 seconds
                    )
                except Exception:
                    # If timeout, we still try to read whatever is there
                    print("  Wait timed out, attempting to capture current value.")

                # 4. Output capture with fallbacks
                output_text = page.evaluate("""() => {
                    const areas = document.querySelectorAll('textarea');
                    // Prefer second textarea, fallback to contenteditable div
                    const target = areas.length > 1 ? areas[1] : document.querySelector('div[contenteditable="true"]');
                    if (!target) return "";
                    return (target.value || target.innerText || "").trim();
                }""")
                
                # Update previous output tracker
                prev_output = output_text

                # 5. Debugging
                print("DEBUG OUTPUT:", output_text)

                # 6. Results Comparison and Excel Update
                actual_val = str(output_text).strip()
                expected_val = str(expected_output).strip() if expected_output else ""
                status = "Pass" if actual_val == expected_val else "Fail"
                
                ws.cell(row=row, column=actual_col).value = output_text
                ws.cell(row=row, column=status_col).value = status
                print(f"  -> Status: {status}")

                # 7. Save Workbook after each row
                wb.save(EXCEL_FILE)
                
            except Exception as e:
                print(f"  Error processing row {row}: {e}")
                ws.cell(row=row, column=status_col).value = "Error"
                wb.save(EXCEL_FILE)

        print("\nTesting complete. Results saved.")
        browser.close()

if __name__ == "__main__":
    run_automation()
