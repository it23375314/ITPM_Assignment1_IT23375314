import openpyxl
import os

file_path = "Assignment 1 - Test cases.xlsx"
if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path)
    print(f"Sheets: {wb.sheetnames}")
    ws = wb.active
    print(f"Active Sheet: {ws.title}")
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
else:
    print(f"File {file_path} not found.")
