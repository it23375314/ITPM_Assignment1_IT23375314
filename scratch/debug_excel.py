import openpyxl
wb = openpyxl.load_workbook("Assignment 1 - Test cases.xlsx")
ws = wb[" Test cases"]
print(f"Max row: {ws.max_row}")
for row in ws.iter_rows(values_only=True):
    print(row)
