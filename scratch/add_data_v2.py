import openpyxl
wb = openpyxl.load_workbook("Assignment 1 - Test cases.xlsx")
ws = wb[" Test cases"]
ws.cell(row=2, column=3).value = "oyata kohomada"  # C2
ws.cell(row=2, column=4).value = "ඔයාට කොහොමද"      # D2
ws.cell(row=6, column=3).value = "mama yanawa"      # C6
ws.cell(row=6, column=4).value = "මම යනවා"          # D6
wb.save("Assignment 1 - Test cases.xlsx")
print("Added sample data to top-left of merged cells.")
