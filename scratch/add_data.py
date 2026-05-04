import openpyxl
wb = openpyxl.load_workbook("Assignment 1 - Test cases.xlsx")
ws = wb[" Test cases"]
ws.cell(row=2, column=3).value = "oyata kohomada"  # Input
ws.cell(row=2, column=4).value = "ඔයාට කොහොමද"      # Expected
ws.cell(row=3, column=3).value = "mama yanawa"      # Input
ws.cell(row=3, column=4).value = "මම යනවා"          # Expected
wb.save("Assignment 1 - Test cases.xlsx")
print("Added sample data.")
